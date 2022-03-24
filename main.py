from turtle import back
from pip import main
import openpyxl
from matplotlib.pyplot import get, text
from numpy import append
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

####
#This will read in the data from the provided xlsx and populate a new workbook with the desired parameters
####

#Initializing Driver
#Passing the "options" argument to supress browser windows
PATH = "C:\Program Files (x86)\chromedriver.exe"
op = webdriver.ChromeOptions()
op.add_argument('headless')
driver = webdriver.Chrome(PATH, options=op)
driver.get("https://www.zillow.com/")



#Initializing xlsx workbooks and sheets
wbRead = load_workbook('FILE 1 LOS ANGELES CA.xlsx')
wbWrite = Workbook()

wsRead = wbRead.active
wsWrite = wbWrite.active
wsWrite.title = "Zillow Data"

#Naming the columns with respect to the desired metrics
wsWrite.append(['Zip Code', 'Zestimate', 'Type', 'Year Built',
'Square Feet'])


#Finding the search bar
search_bar = driver.find_element(By.ID, 'search-box-input')

#Get search terms
#Load and search
for row in range (2,329904):
    search_term = ""
    for col in range(2,5):
        char = get_column_letter(col)
        search_term += wsRead[char + str(row)].value() + " "

    appndList = []

    search_bar.clear()
    search_bar.send_keys(search_term)
    search_bar.send_keys(Keys.RETURN)

    #Capturing information for each parameter and appending a list to the workbook
    try:
        #Zip Code
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Text-c11n-8-62-5__sc-aiai24-0 kZKvMY"))))
            zip_code = element.text()
            appndList.append(zip_code[:-5].encode('utf-8'))
        except:
            appndList.append("N/A")

        #Zestimate
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Text-c11n-8-62-5__sc-aiai24-0 eqgHgX"))))
            priceTemp = element.text().encode('utf-8')
            price = int("".join(filter(str.isdigit, priceTemp)))
            appndList.append(price)
        except:
            appndList.append("N/A")

        #Type
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Text-c11n-8-62-5__sc-aiai24-0 dpf__sc-2arhs5-3 kZKvMY btxEYg"))))
            type = element.text()
            appndList.append(type.encode('utf-8'))
        except:
            appndList.append("N/A")

        #Year Built
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Text-c11n-8-62-5__sc-aiai24-0 dpf__sc-2arhs5-3 kZKvMY btxEYg"))))
            year = element.text()
            appndList.append(year.encode('utf-8'))
        except:
            appndList.append("N/A")

        #Square Feet
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Text-c11n-8-62-5__sc-aiai24-0 kZKvMY"))))
            sFeetTemp = element.text().encode('utf-8')
            squareFeet = int("".join(filter(str.isdigit, sFeetTemp)))
            appndList.append(squareFeet)
        except:
            appndList.append("N/A")

    #if these elements can't be fetched and updated properly then we will go back and start the loop for the next search term       
    except:
        driver.back()
    
    wsWrite.append(appndList)


driver.quit

wbWrite.save('zillow_data.xlsx')   
